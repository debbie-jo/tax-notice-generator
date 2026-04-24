#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
종합소득세 안내문 생성기 - 샘플 Excel 파일 생성
openpyxl을 사용하여 완벽한 스타일과 드롭다운이 적용된 Excel 파일 생성
"""

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import datetime

def create_sample_excel():
    """샘플 Excel 파일 생성"""
    
    # 워크북 생성
    wb = Workbook()
    ws = wb.active
    ws.title = '종소세_입력명단'
    
    # 색상 정의 (RGB)
    colors = {
        'header': 'B8DDE7',      # 라이트 블루 (헤더)
        'headerMain': 'DAEEF3',  # 라이트 블루 (컬럼명)
        'headerSub': 'E6F5F8',   # 매우 라이트 블루 (설명)
        'income': 'FFF9E6',      # 연노랑 (소득유형)
        'deduction': 'F0FFF0',   # 연초록 (공제/서류)
        'sincere': 'F3EEFF',     # 연보라 (성실신고)
        'newRow': 'E8F5E9',      # 연초록 (신규)
        'border': 'A8D4DC'       # 테두리 색
    }
    
    # 테두리 스타일
    thin_border = Border(
        left=Side(style='thin', color=colors['border']),
        right=Side(style='thin', color=colors['border']),
        top=Side(style='thin', color=colors['border']),
        bottom=Side(style='thin', color=colors['border'])
    )
    
    def cell_style(bg_color, bold=False, font_size=11, wrap=False, align='center'):
        """셀 스타일 함수"""
        return {
            'fill': PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid'),
            'font': Font(name='Noto Sans KR', size=font_size, bold=bold, color='000000'),
            'alignment': Alignment(horizontal=align, vertical='center', wrap_text=wrap),
            'border': thin_border
        }
    
    # 열 너비 설정
    column_widths = [5, 22, 10, 8, 8, 8, 8, 9, 8, 13, 12, 10, 10, 12, 20]
    for idx, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    
    # 행 높이 설정
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 42
    ws.row_dimensions[3].height = 18
    for r in range(4, 7):
        ws.row_dimensions[r].height = 18
    
    # ═══ 행 1: 그룹 헤더 ═══
    headers_group = [
        (1, 1, 4, '기본 정보 (자동)'),
        (1, 5, 9, '소득 유형  ※ 해당되면 O 입력'),
        (1, 10, 12, '차량 정보'),
        (1, 13, 13, '성실신고'),
        (1, 14, 14, '진행 체크'),
        (1, 15, 15, '비고')
    ]
    
    for r, c_start, c_end, text in headers_group:
        cell = ws.cell(row=r, column=c_start)
        cell.value = text
        style = cell_style(colors['header'], bold=True, wrap=True)
        cell.fill = style['fill']
        cell.font = style['font']
        cell.alignment = style['alignment']
        cell.border = style['border']
        
        # 병합
        if c_start < c_end:
            ws.merge_cells(f'{get_column_letter(c_start)}{r}:{get_column_letter(c_end)}{r}')
    
    # ═══ 행 2: 컬럼명 ═══
    headers = [
        'No', '업체명', '대표자명', '상태',
        '사업\n소득', '근로\n소득', '임대\n소득', '이자·\n배당', '기타\n소득',
        '차량번호', '차량유형\n(소유/렌트/리스)', '보험료증권\n필요',
        '성실신고\n대상', '서류\n수취', '비고'
    ]
    
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_idx)
        cell.value = header
        style = cell_style(colors['headerMain'], bold=True, wrap=True)
        cell.fill = style['fill']
        cell.font = style['font']
        cell.alignment = style['alignment']
        cell.border = style['border']
    
    # ═══ 행 3: 설명 힌트 ═══
    hints = [
        '', '', '', '',
        '복수 가능', '복수 가능', '복수 가능', '복수 가능', '복수 가능',
        '예) 12가3456', '소유/렌트/리스', '차량번호\n입력시 자동',
        'O 입력', '완료/미완료/불필요', '직접 입력'
    ]
    
    for col_idx, hint in enumerate(hints, 1):
        cell = ws.cell(row=3, column=col_idx)
        cell.value = hint
        style = cell_style(colors['headerSub'], font_size=8, wrap=True)
        style['font'] = Font(name='Noto Sans KR', size=8, color='777777')
        cell.fill = style['fill']
        cell.font = style['font']
        cell.alignment = style['alignment']
        cell.border = style['border']
    
    # ═══ 샘플 데이터 (행 4~6) ═══
    samples = [
        {
            'bg': colors['newRow'],
            'no': 1,
            'vals': ['홍길동식당', '홍길동', '활성', 'O', 'O', '', '', '', '', '소유', '', '', '', '근로+사업']
        },
        {
            'bg': 'FFFFFF',
            'no': 2,
            'vals': ['김영희컨설팅', '김영희', '활성', 'O', '', '', 'O', '', '', '렌트', '', '', '', '이자배당']
        },
        {
            'bg': 'EFF6FF',
            'no': 3,
            'vals': ['이철수사무소', '이철수', '활성', 'O', 'O', '', 'O', 'O', '12가3456', '소유', '', 'O', '', '성실신고+차량']
        }
    ]
    
    for row_idx, sample in enumerate(samples, 4):
        # No 열
        cell = ws.cell(row=row_idx, column=1)
        cell.value = sample['no']
        style = cell_style(sample['bg'])
        cell.fill = style['fill']
        cell.font = style['font']
        cell.alignment = style['alignment']
        cell.border = style['border']
        
        # 데이터 열들
        for col_idx, val in enumerate(sample['vals'], 2):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = val if val else None
            
            # 배경색 구분
            bg = sample['bg']
            if 4 <= col_idx <= 8:  # 소득유형
                bg = colors['income']
            elif col_idx == 10:  # 차량유형
                bg = colors['income']
            elif col_idx == 12:  # 성실신고
                bg = colors['sincere']
            elif col_idx == 13:  # 서류수취
                bg = colors['deduction']
            
            style = cell_style(bg, align='left' if col_idx == 2 else 'center')
            cell.fill = style['fill']
            cell.font = style['font']
            cell.alignment = style['alignment']
            cell.border = style['border']
    
    # ═══ 드롭다운 (데이터 유효성) ═══
    # K열 (차량유형)
    dv_vehicle = DataValidation(
        type='list',
        formula1='"소유,렌트,리스"',
        allow_blank=True,
        showDropDown=True
    )
    dv_vehicle.error = '소유, 렌트, 리스 중 선택하세요'
    dv_vehicle.errorTitle = '입력 오류'
    ws.add_data_validation(dv_vehicle)
    dv_vehicle.add(f'K4:K1000')
    
    # M열 (성실신고)
    dv_sincere = DataValidation(
        type='list',
        formula1='"O"',
        allow_blank=True,
        showDropDown=True
    )
    dv_sincere.error = 'O를 입력하세요'
    dv_sincere.errorTitle = '입력 오류'
    ws.add_data_validation(dv_sincere)
    dv_sincere.add(f'M4:M1000')
    
    # N열 (서류수취)
    dv_status = DataValidation(
        type='list',
        formula1='"완료,미완료,불필요"',
        allow_blank=True,
        showDropDown=True
    )
    dv_status.error = '완료, 미완료, 불필요 중 선택하세요'
    dv_status.errorTitle = '입력 오류'
    ws.add_data_validation(dv_status)
    dv_status.add(f'N4:N1000')
    
    # ═══ 틀 고정 ═══
    ws.freeze_panes = 'E4'
    
    # ═══ 파일 저장 ═══
    filename = f'종소세_입력명단_샘플_{datetime.now().strftime("%Y%m%d")}.xlsx'
    wb.save(filename)
    print(f'✅ Excel 파일 생성 완료: {filename}')
    return filename

if __name__ == '__main__':
    create_sample_excel()
